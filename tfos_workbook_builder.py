"""
TFOS v0.1 Workbook Generator
Thome Farm Operating System - Excel Architecture

This script generates the complete TFOS v0.1 workbook structure with
all required worksheets and structured tables.

Technology: openpyxl
Standards: Clean, documented, production-quality code
"""

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class TFOSWorkbookBuilder:
    """Builder class for TFOS v0.1 workbook structure."""
    
    def __init__(self, filename="TFOS_v0.1.xlsx"):
        """Initialize workbook with professional formatting standards."""
        self.wb = Workbook()
        self.filename = filename
        self.ws_names = [
            "README",
            "Settings",
            "Fields",
            "Equipment",
            "Loans",
            "Family Financials",
            "Crop Budgets",
            "Operating Costs",
            "Harvest Import",
            "Planting Import",
            "Application Import",
            "Loan Amortization",
            "Revenue by Field",
            "Cost by Field",
            "Field Profitability",
            "Balance Sheet",
            "Cash Flow",
            "Dashboard",
            "Developer Notes"
        ]
        
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
    
    def _get_header_fill(self):
        """Professional header fill color."""
        return PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    def _get_header_font(self):
        """Professional header font."""
        return Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    def _get_border(self):
        """Professional border style."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        return thin_border
    
    def _create_table(self, worksheet, table_name, headers, start_row=1):
        """
        Create a structured Excel table with headers.
        
        Args:
            worksheet: Target worksheet
            table_name: Name of the table
            headers: List of column headers
            start_row: Starting row for table (default 1)
        """
        # Add headers
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=start_row, column=col_idx, value=header)
            cell.fill = self._get_header_fill()
            cell.font = self._get_header_font()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self._get_border()
        
        # Create table
        end_col = get_column_letter(len(headers))
        table_ref = f"A{start_row}:{end_col}{start_row + 1}"
        
        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        worksheet.add_table(tab)
        
        # Set column widths
        for col_idx in range(1, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 18
        
        # Set row height for headers
        worksheet.row_dimensions[start_row].height = 25
        
        return table_ref
    
    def create_readme(self):
        """Create README worksheet with project information."""
        ws = self.wb.create_sheet("README", 0)
        ws.sheet_properties.tabColor = "FF0000"
        
        # Title
        title_cell = ws["A1"]
        title_cell.value = "TFOS v0.1 - Thome Farm Operating System"
        title_cell.font = Font(name="Calibri", size=16, bold=True)
        ws.row_dimensions[1].height = 30
        
        # Subtitle
        ws["A2"] = "Financial Operating System for Row Crop Farming"
        ws["A2"].font = Font(name="Calibri", size=12, italic=True)
        ws.row_dimensions[2].height = 20
        
        # Generation info
        ws["A4"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A4"].font = Font(name="Calibri", size=10, italic=True, color="666666")
        
        # Feature list
        ws["A6"] = "v0.1 Capabilities:"
        ws["A6"].font = Font(name="Calibri", size=11, bold=True)
        
        features = [
            "Farm Financial Statements",
            "Field Profitability Analysis",
            "Equipment Tracking",
            "Loan Amortization",
            "Crop Budgets",
            "John Deere Operations Center Imports",
            "Executive Dashboard"
        ]
        
        for idx, feature in enumerate(features, start=7):
            ws[f"A{idx}"] = f"• {feature}"
            ws[f"A{idx}"].font = Font(name="Calibri", size=10)
        
        # Worksheets guide
        ws[f"A{7 + len(features) + 2}"] = "Worksheets:"
        ws[f"A{7 + len(features) + 2}"].font = Font(name="Calibri", size=11, bold=True)
        
        descriptions = {
            "Settings": "Farm operating parameters and assumptions",
            "Fields": "Field master data and characteristics",
            "Equipment": "Equipment inventory and specifications",
            "Loans": "Loan principal and terms",
            "Family Financials": "Personal/family financial data",
            "Crop Budgets": "Crop-by-crop budget planning",
            "Operating Costs": "Operational expense categories",
            "Harvest Import": "John Deere Operations Center harvest data",
            "Planting Import": "John Deere Operations Center planting data",
            "Application Import": "John Deere Operations Center application data",
            "Loan Amortization": "Loan payment schedules",
            "Revenue by Field": "Field-level revenue analysis",
            "Cost by Field": "Field-level cost analysis",
            "Field Profitability": "Field profitability calculations",
            "Balance Sheet": "Farm balance sheet",
            "Cash Flow": "Farm cash flow statement",
            "Dashboard": "Executive summary dashboard",
            "Developer Notes": "Development documentation"
        }
        
        current_row = 7 + len(features) + 3
        for ws_name, description in descriptions.items():
            ws[f"A{current_row}"] = f"  {ws_name}"
            ws[f"B{current_row}"] = description
            ws[f"A{current_row}"].font = Font(name="Calibri", size=10)
            ws[f"B{current_row}"].font = Font(name="Calibri", size=10, italic=True, color="444444")
            current_row += 1
        
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50
    
    def create_settings(self):
        """Create Settings worksheet - centralized configuration."""
        ws = self.wb.create_sheet("Settings")
        
        headers = ["Setting Name", "Value", "Description", "Data Type"]
        self._create_table(ws, "SettingsTable", headers)
        
        # Add template rows for settings structure
        settings_data = [
            ("Farm Name", "", "Legal farm business name", "Text"),
            ("Tax Year", "", "Current tax year (YYYY)", "Number"),
            ("Farm ID", "", "Unique farm identifier", "Text"),
            ("Currency", "USD", "Operating currency", "Text"),
            ("Default Crop Year", "", "Current crop year (YYYY)", "Number"),
            ("Data Import Source", "John Deere Operations Center", "Source of operational data", "Text"),
            ("Discount Rate", "", "Used for NPV calculations", "Percentage"),
            ("Inflation Rate", "", "Annual inflation assumption", "Percentage"),
        ]
        
        for idx, (name, value, desc, dtype) in enumerate(settings_data, start=2):
            ws[f"A{idx}"] = name
            ws[f"B{idx}"] = value
            ws[f"C{idx}"] = desc
            ws[f"D{idx}"] = dtype
    
    def create_fields(self):
        """Create Fields worksheet - field master data."""
        ws = self.wb.create_sheet("Fields")
        
        headers = [
            "Field ID",
            "Field Name",
            "Acres",
            "Crop",
            "Soil Type",
            "GPS Coordinates",
            "Tillage Practice",
            "Previous Crop",
            "Organic",
            "Notes"
        ]
        self._create_table(ws, "FieldsTable", headers)
    
    def create_equipment(self):
        """Create Equipment worksheet - equipment inventory."""
        ws = self.wb.create_sheet("Equipment")
        
        headers = [
            "Equipment ID",
            "Equipment Name",
            "Equipment Type",
            "Manufacturer",
            "Model",
            "Serial Number",
            "Year",
            "Hours/Miles",
            "Acquisition Date",
            "Acquisition Cost",
            "Current Value",
            "Status",
            "Notes"
        ]
        self._create_table(ws, "EquipmentTable", headers)
    
    def create_loans(self):
        """Create Loans worksheet - loan principal and terms."""
        ws = self.wb.create_sheet("Loans")
        
        headers = [
            "Loan ID",
            "Lender",
            "Loan Type",
            "Principal Amount",
            "Interest Rate",
            "Origination Date",
            "Maturity Date",
            "Payment Frequency",
            "Annual Payments",
            "Term (Years)",
            "Status",
            "Notes"
        ]
        self._create_table(ws, "LoansTable", headers)
    
    def create_family_financials(self):
        """Create Family Financials worksheet - personal financial data."""
        ws = self.wb.create_sheet("Family Financials")
        
        headers = [
            "Category",
            "Person Name",
            "Income Type",
            "Annual Amount",
            "Expense Type",
            "Notes"
        ]
        self._create_table(ws, "FamilyFinancialsTable", headers)
    
    def create_crop_budgets(self):
        """Create Crop Budgets worksheet - crop-by-crop planning."""
        ws = self.wb.create_sheet("Crop Budgets")
        
        headers = [
            "Budget ID",
            "Crop",
            "Crop Year",
            "Field ID",
            "Acres",
            "Budget Category",
            "Item Description",
            "Unit",
            "Quantity",
            "Unit Cost",
            "Total Cost",
            "Notes"
        ]
        self._create_table(ws, "CropBudgetsTable", headers)
    
    def create_operating_costs(self):
        """Create Operating Costs worksheet - expense categories."""
        ws = self.wb.create_sheet("Operating Costs")
        
        headers = [
            "Cost ID",
            "Cost Category",
            "Cost Type",
            "Description",
            "Amount",
            "Date Incurred",
            "Field ID",
            "Equipment ID",
            "Vendor",
            "Invoice Number",
            "Notes"
        ]
        self._create_table(ws, "OperatingCostsTable", headers)
    
    def create_harvest_import(self):
        """Create Harvest Import worksheet - JD Operations Center harvest data."""
        ws = self.wb.create_sheet("Harvest Import")
        
        headers = [
            "Import ID",
            "Field ID",
            "Crop",
            "Harvest Date",
            "Equipment ID",
            "Total Bushels",
            "Bushels per Acre",
            "Moisture",
            "Test Weight",
            "Damage",
            "Foreign Material",
            "Data Source",
            "Import Date",
            "Notes"
        ]
        self._create_table(ws, "HarvestImportTable", headers)
    
    def create_planting_import(self):
        """Create Planting Import worksheet - JD Operations Center planting data."""
        ws = self.wb.create_sheet("Planting Import")
        
        headers = [
            "Import ID",
            "Field ID",
            "Crop",
            "Planting Date",
            "Equipment ID",
            "Acres Planted",
            "Seed Variety",
            "Seed Rate",
            "Target Population",
            "Hybrid/Variety Notes",
            "Data Source",
            "Import Date",
            "Notes"
        ]
        self._create_table(ws, "PlantingImportTable", headers)
    
    def create_application_import(self):
        """Create Application Import worksheet - JD Operations Center application data."""
        ws = self.wb.create_sheet("Application Import")
        
        headers = [
            "Import ID",
            "Field ID",
            "Crop",
            "Application Date",
            "Equipment ID",
            "Application Type",
            "Product Name",
            "Rate Applied",
            "Rate Unit",
            "Acres Applied",
            "Cost per Acre",
            "Total Cost",
            "Data Source",
            "Import Date",
            "Notes"
        ]
        self._create_table(ws, "ApplicationImportTable", headers)
    
    def create_loan_amortization(self):
        """Create Loan Amortization worksheet - loan payment schedules."""
        ws = self.wb.create_sheet("Loan Amortization")
        
        headers = [
            "Payment Number",
            "Loan ID",
            "Payment Date",
            "Beginning Balance",
            "Payment Amount",
            "Principal Payment",
            "Interest Payment",
            "Ending Balance",
            "Status"
        ]
        self._create_table(ws, "LoanAmortizationTable", headers)
    
    def create_revenue_by_field(self):
        """Create Revenue by Field worksheet - field-level revenue analysis."""
        ws = self.wb.create_sheet("Revenue by Field")
        
        headers = [
            "Field ID",
            "Crop Year",
            "Crop",
            "Bushels/Units Produced",
            "Yield per Acre",
            "Price per Unit",
            "Gross Revenue",
            "Government Payments",
            "Insurance Proceeds",
            "Total Revenue",
            "Notes"
        ]
        self._create_table(ws, "RevenueByFieldTable", headers)
    
    def create_cost_by_field(self):
        """Create Cost by Field worksheet - field-level cost analysis."""
        ws = self.wb.create_sheet("Cost by Field")
        
        headers = [
            "Field ID",
            "Crop Year",
            "Crop",
            "Seed Cost",
            "Fertilizer Cost",
            "Crop Protection Cost",
            "Fuel and Lube Cost",
            "Labor Cost",
            "Equipment Cost",
            "Land Cost",
            "Overhead Allocation",
            "Total Cost",
            "Cost per Acre",
            "Notes"
        ]
        self._create_table(ws, "CostByFieldTable", headers)
    
    def create_field_profitability(self):
        """Create Field Profitability worksheet - profitability calculations."""
        ws = self.wb.create_sheet("Field Profitability")
        
        headers = [
            "Field ID",
            "Crop Year",
            "Crop",
            "Acres",
            "Total Revenue",
            "Total Cost",
            "Gross Profit",
            "Gross Margin %",
            "Revenue per Acre",
            "Cost per Acre",
            "Profit per Acre",
            "ROI %",
            "Notes"
        ]
        self._create_table(ws, "FieldProfitabilityTable", headers)
    
    def create_balance_sheet(self):
        """Create Balance Sheet worksheet - farm balance sheet."""
        ws = self.wb.create_sheet("Balance Sheet")
        
        headers = [
            "Line Item",
            "Category",
            "Subcategory",
            "As of Date",
            "Amount",
            "Notes"
        ]
        self._create_table(ws, "BalanceSheetTable", headers)
        
        # Add section structure
        sections = [
            ("ASSETS", "Current Assets", "Cash and Cash Equivalents"),
            ("ASSETS", "Current Assets", "Accounts Receivable"),
            ("ASSETS", "Current Assets", "Inventory"),
            ("ASSETS", "Fixed Assets", "Land"),
            ("ASSETS", "Fixed Assets", "Equipment"),
            ("ASSETS", "Fixed Assets", "Accumulated Depreciation"),
            ("LIABILITIES", "Current Liabilities", "Accounts Payable"),
            ("LIABILITIES", "Current Liabilities", "Current Portion of Debt"),
            ("LIABILITIES", "Long-Term Liabilities", "Long-Term Debt"),
            ("EQUITY", "Owners Equity", "Beginning Equity"),
            ("EQUITY", "Owners Equity", "Current Year Income"),
        ]
    
    def create_cash_flow(self):
        """Create Cash Flow worksheet - farm cash flow statement."""
        ws = self.wb.create_sheet("Cash Flow")
        
        headers = [
            "Line Item",
            "Category",
            "Period",
            "Amount",
            "Cumulative",
            "Notes"
        ]
        self._create_table(ws, "CashFlowTable", headers)
    
    def create_dashboard(self):
        """Create Dashboard worksheet - executive summary."""
        ws = self.wb.create_sheet("Dashboard")
        ws.sheet_properties.tabColor = "00B050"
        
        # Title
        title_cell = ws["A1"]
        title_cell.value = "TFOS Executive Dashboard"
        title_cell.font = Font(name="Calibri", size=14, bold=True)
        ws.row_dimensions[1].height = 25
        
        # Subtitle with date
        ws["A2"] = "Farm Financial Summary"
        ws["A2"].font = Font(name="Calibri", size=11)
        
        # Create summary metrics table
        headers = [
            "Metric",
            "Value",
            "Target",
            "Variance",
            "Status"
        ]
        self._create_table(ws, "DashboardMetricsTable", headers, start_row=4)
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 15
    
    def create_developer_notes(self):
        """Create Developer Notes worksheet - development documentation."""
        ws = self.wb.create_sheet("Developer Notes")
        ws.sheet_properties.tabColor = "0070C0"
        
        # Title
        title_cell = ws["A1"]
        title_cell.value = "TFOS v0.1 Developer Notes"
        title_cell.font = Font(name="Calibri", size=14, bold=True)
        ws.row_dimensions[1].height = 25
        
        # Notes table
        headers = [
            "Date",
            "Developer",
            "Category",
            "Status",
            "Description",
            "Notes"
        ]
        self._create_table(ws, "DeveloperNotesTable", headers, start_row=3)
        
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 40
        ws.column_dimensions["F"].width = 35
    
    def build(self):
        """Build complete workbook with all worksheets."""
        print("Building TFOS v0.1 Workbook Structure...")
        
        self.create_readme()
        self.create_settings()
        self.create_fields()
        self.create_equipment()
        self.create_loans()
        self.create_family_financials()
        self.create_crop_budgets()
        self.create_operating_costs()
        self.create_harvest_import()
        self.create_planting_import()
        self.create_application_import()
        self.create_loan_amortization()
        self.create_revenue_by_field()
        self.create_cost_by_field()
        self.create_field_profitability()
        self.create_balance_sheet()
        self.create_cash_flow()
        self.create_dashboard()
        self.create_developer_notes()
        
        print(f"✓ All {len(self.ws_names)} worksheets created")
        return self.wb
    
    def save(self):
        """Save workbook to file."""
        self.build()
        self.wb.save(self.filename)
        print(f"✓ Workbook saved: {self.filename}")
        return self.filename


if __name__ == "__main__":
    builder = TFOSWorkbookBuilder("TFOS_v0.1.xlsx")
    builder.save()
    print("\nTFOS v0.1 workbook structure complete.")
    print("Ready for formula development and data imports.")
