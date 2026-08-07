"""
TFOS v0.1 Settings Worksheet Builder
Thome Farm Operating System - Centralized Configuration

Creates the Settings worksheet with all editable assumptions organized by category.
Every value includes named ranges for formula references.

Standards:
- Clean, documented code
- All assumptions centralized
- Named ranges for all values
- Professional formatting
- Categories and descriptions
"""

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class SettingsBuilder:
    """Builder for TFOS Settings worksheet with centralized assumptions."""
    
    def __init__(self):
        """Initialize workbook and styling."""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Settings"
        self.named_ranges = []
        
        # Color scheme
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.category_font = Font(name="Calibri", size=11, bold=True, color="1F4E78")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _create_header(self):
        """Create worksheet header and instructions."""
        # Title
        title = self.ws["A1"]
        title.value = "TFOS Settings - Centralized Assumptions"
        title.font = Font(name="Calibri", size=14, bold=True)
        title.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[1].height = 25
        
        # Generated date
        date_cell = self.ws["A2"]
        date_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
        
        # Instructions
        instructions_row = 4
        self.ws[f"A{instructions_row}"] = "IMPORTANT: All editable assumptions are stored here."
        self.ws[f"A{instructions_row}"].font = Font(name="Calibri", size=10, bold=True)
        
        self.ws[f"A{instructions_row + 1}"] = "Other worksheets reference these values using named ranges."
        self.ws[f"A{instructions_row + 1}"].font = Font(name="Calibri", size=10)
        
        self.ws[f"A{instructions_row + 2}"] = "Do not delete rows or change column structure."
        self.ws[f"A{instructions_row + 2}"].font = Font(name="Calibri", size=10, color="FF0000")
        
        return instructions_row + 4
    
    def _format_header_row(self, row_num, headers):
        """Format column headers."""
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=row_num, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.border
        
        self.ws.row_dimensions[row_num].height = 30
    
    def _format_category_row(self, row_num):
        """Format category separator row."""
        for col_idx in range(1, 6):
            cell = self.ws.cell(row=row_num, column=col_idx)
            cell.fill = self.category_fill
            cell.font = self.category_font
            cell.border = self.border
        
        self.ws.row_dimensions[row_num].height = 20
    
    def _add_setting(self, row_num, category, setting, value, units, description, create_named_range=True):
        """
        Add a setting row with all columns.
        
        Args:
            row_num: Row number
            category: Setting category
            setting: Setting name
            value: Current value
            units: Unit of measurement
            description: Setting description
            create_named_range: Whether to create a named range for this value
        """
        # Column A: Category
        cell_a = self.ws.cell(row=row_num, column=1, value=category)
        cell_a.alignment = Alignment(horizontal="left", vertical="center")
        cell_a.border = self.border
        cell_a.font = Font(name="Calibri", size=10)
        
        # Column B: Setting
        cell_b = self.ws.cell(row=row_num, column=2, value=setting)
        cell_b.alignment = Alignment(horizontal="left", vertical="center")
        cell_b.border = self.border
        cell_b.font = Font(name="Calibri", size=10, bold=True)
        
        # Column C: Current Value
        cell_c = self.ws.cell(row=row_num, column=3, value=value)
        cell_c.alignment = Alignment(horizontal="right", vertical="center")
        cell_c.border = self.border
        cell_c.font = Font(name="Calibri", size=10)
        
        # Column D: Units
        cell_d = self.ws.cell(row=row_num, column=4, value=units)
        cell_d.alignment = Alignment(horizontal="left", vertical="center")
        cell_d.border = self.border
        cell_d.font = Font(name="Calibri", size=10, color="666666")
        
        # Column E: Description
        cell_e = self.ws.cell(row=row_num, column=5, value=description)
        cell_e.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_e.border = self.border
        cell_e.font = Font(name="Calibri", size=10, italic=True)
        
        # Create named range if requested
        if create_named_range:
            range_name = f"{setting.replace(' ', '_').replace('-', '_').upper()}"
            try:
                self.wb.named_ranges[range_name] = f"Settings!$C${row_num}"
                self.named_ranges.append({
                    'name': range_name,
                    'cell': f"C{row_num}",
                    'setting': setting
                })
            except:
                pass  # Handle duplicate names gracefully
        
        self.ws.row_dimensions[row_num].height = 18
        return row_num + 1
    
    def build_farm_section(self, current_row):
        """Build FARM category settings."""
        # Section header
        self.ws[f"A{current_row}"] = "FARM"
        self._format_category_row(current_row)
        current_row += 1
        
        farm_settings = [
            ("Farm Name", "", "Text", "Legal farm business name"),
            ("Tax Year", 2026, "Year", "Current tax year for financial statements"),
            ("Farm ID", "", "Text", "Unique identifier for this farm operation"),
            ("Primary Crop", "Corn", "Text", "Main crop produced"),
            ("Secondary Crop", "Soybeans", "Text", "Secondary crop produced"),
            ("Total Acres", 0, "Acres", "Total farm size"),
            ("Owned Acres", 0, "Acres", "Acres owned by operator"),
            ("Leased Acres", 0, "Acres", "Acres leased or rented"),
            ("Crop Rotation", "Corn-Soybeans", "Text", "Rotation pattern used"),
            ("Farm Status", "Active", "Text", "Active, Transition, or Inactive"),
        ]
        
        for setting, value, units, desc in farm_settings:
            current_row = self._add_setting(current_row, "Farm", setting, value, units, desc)
        
        return current_row + 1
    
    def build_financial_section(self, current_row):
        """Build FINANCIAL category settings."""
        # Section header
        self.ws[f"A{current_row}"] = "FINANCIAL"
        self._format_category_row(current_row)
        current_row += 1
        
        financial_settings = [
            ("Discount Rate", 0.05, "Decimal (5% = 0.05)", "Used for NPV and time value of money calculations"),
            ("Target Debt-to-Asset Ratio", 0.40, "Decimal (40% = 0.40)", "Maximum acceptable leverage ratio"),
            ("Operating Capital Requirement", 0, "USD", "Minimum cash reserve required for operations"),
            ("Accounting Method", "Cash", "Text", "Cash or Accrual accounting"),
        ]
        
        for setting, value, units, desc in financial_settings:
            current_row = self._add_setting(current_row, "Financial", setting, value, units, desc)
        
        return current_row + 1
    
    def build_family_section(self, current_row):
        """Build FAMILY category settings."""
        # Section header
        self.ws[f"A{current_row}"] = "FAMILY"
        self._format_category_row(current_row)
        current_row += 1
        
        family_settings = [
            ("Operator Name", "", "Text", "Primary farm operator name"),
            ("Spouse Name", "", "Text", "Spouse name (if applicable)"),
            ("Number of Dependents", 0, "Count", "Number of dependents claimed"),
            ("Family Income Goal", 0, "USD", "Desired annual family income from farm"),
            ("Living Expense Budget", 0, "USD", "Annual family living expenses"),
        ]
        
        for setting, value, units, desc in family_settings:
            current_row = self._add_setting(current_row, "Family", setting, value, units, desc)
        
        return current_row + 1
    
    def build_retirement_section(self, current_row):
        """Build RETIREMENT category settings."""
        # Section header
        self.ws[f"A{current_row}"] = "RETIREMENT"
        self._format_category_row(current_row)
        current_row += 1
        
        retirement_settings = [
            ("Retirement Year", 2040, "Year", "Target year for farm transition/retirement"),
            ("Years to Retirement", 14, "Years", "Years until planned retirement"),
            ("Succession Plan", "Undetermined", "Text", "Undetermined, Family, Sale, or Other"),
            ("Retirement Income Need", 0, "USD/Year", "Desired annual retirement income"),
        ]
        
        for setting, value, units, desc in retirement_settings:
            current_row = self._add_setting(current_row, "Retirement", setting, value, units, desc)
        
        return current_row + 1
    
    def build_inflation_section(self, current_row):
        """Build INFLATION category settings."""
        # Section header
        self.ws[f"A{current_row}"] = "INFLATION"
        self._format_category_row(current_row)
        current_row += 1
        
        inflation_settings = [
            ("General Inflation Rate", 0.025, "Decimal (2.5% = 0.025)", "Overall inflation assumption"),
            ("Input Cost Inflation", 0.030, "Decimal (3.0% = 0.030)", "Inflation on seed, fertilizer, chemicals"),
            ("Fuel Inflation Rate", 0.040, "Decimal (4.0% = 0.040)", "Inflation on fuel and energy"),
            ("Labor Inflation Rate", 0.025, "Decimal (2.5% = 0.025)", "Inflation on labor costs"),
            ("Equipment Inflation Rate", 0.025, "Decimal (2.5% = 0.025)", "Inflation on machinery and equipment"),
        ]
        
        for setting, value, units, desc in inflation_settings:
            current_row = self._add_setting(current_row, "Inflation", setting, value, units, desc)
        
        return current_row + 1
    
    def build_interest_rates_section(self, current_row):
        """Build INTEREST RATES category settings."""
        # Section header
        self.ws[f"A{current_row}"] = "INTEREST RATES"
        self._format_category_row(current_row)
        current_row += 1
        
        interest_settings = [
            ("Operating Loan Rate", 0.065, "Decimal (6.5% = 0.065)", "Interest rate on operating loans"),
            ("Equipment Loan Rate", 0.055, "Decimal (5.5% = 0.055)", "Interest rate on equipment financing"),
            ("Land Loan Rate", 0.050, "Decimal (5.0% = 0.050)", "Interest rate on land mortgages"),
            ("Line of Credit Rate", 0.075, "Decimal (7.5% = 0.075)", "Interest rate on LOC"),
            ("Savings/CD Rate", 0.035, "Decimal (3.5% = 0.035)", "Expected return on savings accounts"),
        ]
        
        for setting, value, units, desc in interest_settings:
            current_row = self._add_setting(current_row, "Interest Rates", setting, value, units, desc)
        
        return current_row + 1
    
    def build_crop_prices_section(self, current_row):
        """Build CROP PRICES category settings."""
        # Section header
        self.ws[f"A{current_row}"] = "CROP PRICES"
        self._format_category_row(current_row)
        current_row += 1
        
        crop_price_settings = [
            ("Corn Price", 0.00, "USD/Bushel", "Expected selling price for corn"),
            ("Soybean Price", 0.00, "USD/Bushel", "Expected selling price for soybeans"),
            ("Wheat Price", 0.00, "USD/Bushel", "Expected selling price for wheat"),
            ("Hay Price", 0.00, "USD/Ton", "Expected selling price for hay"),
        ]
        
        for setting, value, units, desc in crop_price_settings:
            current_row = self._add_setting(current_row, "Crop Prices", setting, value, units, desc)
        
        return current_row + 1
    
    def build_average_yields_section(self, current_row):
        """Build AVERAGE YIELDS category settings."""
        # Section header
        self.ws[f"A{current_row}"] = "AVERAGE YIELDS"
        self._format_category_row(current_row)
        current_row += 1
        
        yield_settings = [
            ("Corn Yield", 0.00, "Bushels/Acre", "Expected corn yield by field"),
            ("Soybean Yield", 0.00, "Bushels/Acre", "Expected soybean yield by field"),
            ("Wheat Yield", 0.00, "Bushels/Acre", "Expected wheat yield by field"),
            ("Hay Yield", 0.00, "Tons/Acre", "Expected hay yield per acre"),
        ]
        
        for setting, value, units, desc in yield_settings:
            current_row = self._add_setting(current_row, "Average Yields", setting, value, units, desc)
        
        return current_row + 1
    
    def build_machinery_costs_section(self, current_row):
        """Build MACHINERY COSTS category settings."""
        # Section header
        self.ws[f"A{current_row}"] = "MACHINERY COSTS"
        self._format_category_row(current_row)
        current_row += 1
        
        machinery_settings = [
            ("Machinery Repair Rate", 0.04, "Decimal (4% = 0.04)", "Annual repair as % of machinery value"),
            ("Machinery Depreciation Years", 10, "Years", "Average useful life of machinery"),
            ("Depreciation Method", "Straight-Line", "Text", "Straight-Line or Accelerated"),
            ("Annual Machinery Replacement Budget", 0, "USD", "Budget for new machinery purchases"),
            ("Tire Replacement Cycle", 3, "Years", "Years between tire replacements"),
        ]
        
        for setting, value, units, desc in machinery_settings:
            current_row = self._add_setting(current_row, "Machinery Costs", setting, value, units, desc)
        
        return current_row + 1
    
    def build_fuel_section(self, current_row):
        """Build FUEL category settings."""
        # Section header
        self.ws[f"A{current_row}"] = "FUEL"
        self._format_category_row(current_row)
        current_row += 1
        
        fuel_settings = [
            ("Diesel Price", 0.00, "USD/Gallon", "Expected diesel fuel price"),
            ("Gasoline Price", 0.00, "USD/Gallon", "Expected gasoline price"),
            ("Propane Price", 0.00, "USD/Gallon", "Expected propane price"),
            ("Average Fuel Efficiency", 0.00, "Gallons/Hour", "Average fuel consumption for machinery"),
        ]
        
        for setting, value, units, desc in fuel_settings:
            current_row = self._add_setting(current_row, "Fuel", setting, value, units, desc)
        
        return current_row + 1
    
    def build_insurance_section(self, current_row):
        """Build INSURANCE category settings."""
        # Section header
        self.ws[f"A{current_row}"] = "INSURANCE"
        self._format_category_row(current_row)
        current_row += 1
        
        insurance_settings = [
            ("Crop Insurance Coverage", 0.75, "Decimal (75% = 0.75)", "% of expected revenue covered"),
            ("Crop Insurance Cost", 0.00, "USD/Acre", "Annual crop insurance cost per acre"),
            ("Property Insurance Rate", 0.005, "Decimal (0.5% = 0.005)", "Annual property insurance as % of value"),
            ("Liability Insurance Annual", 0, "USD/Year", "Annual liability insurance cost"),
            ("Equipment Insurance Annual", 0, "USD/Year", "Annual equipment insurance cost"),
        ]
        
        for setting, value, units, desc in insurance_settings:
            current_row = self._add_setting(current_row, "Insurance", setting, value, units, desc)
        
        return current_row + 1
    
    def build_tax_rates_section(self, current_row):
        """Build TAX RATES category settings."""
        # Section header
        self.ws[f"A{current_row}"] = "TAX RATES"
        self._format_category_row(current_row)
        current_row += 1
        
        tax_settings = [
            ("Federal Income Tax Rate", 0.22, "Decimal (22% = 0.22)", "Marginal federal income tax rate"),
            ("State Income Tax Rate", 0.05, "Decimal (5% = 0.05)", "Marginal state income tax rate"),
            ("Self-Employment Tax Rate", 0.9235, "Decimal (92.35% = 0.9235)", "Self-employment tax on net earnings"),
            ("Capital Gains Tax Rate", 0.15, "Decimal (15% = 0.15)", "Long-term capital gains tax rate"),
            ("Property Tax Rate", 0.01, "Decimal (1% = 0.01)", "Annual property tax on land value"),
        ]
        
        for setting, value, units, desc in tax_settings:
            current_row = self._add_setting(current_row, "Tax Rates", setting, value, units, desc)
        
        return current_row + 1
    
    def build(self):
        """Build complete Settings worksheet."""
        print("Building TFOS Settings Worksheet...")
        
        # Set column widths
        self.ws.column_dimensions["A"].width = 20
        self.ws.column_dimensions["B"].width = 30
        self.ws.column_dimensions["C"].width = 18
        self.ws.column_dimensions["D"].width = 25
        self.ws.column_dimensions["E"].width = 45
        
        # Create header
        current_row = self._create_header()
        
        # Add header row for table
        headers = ["Category", "Setting", "Current Value", "Units", "Description"]
        self._format_header_row(current_row, headers)
        current_row += 1
        
        # Build all sections
        current_row = self.build_farm_section(current_row)
        current_row = self.build_financial_section(current_row)
        current_row = self.build_family_section(current_row)
        current_row = self.build_retirement_section(current_row)
        current_row = self.build_inflation_section(current_row)
        current_row = self.build_interest_rates_section(current_row)
        current_row = self.build_crop_prices_section(current_row)
        current_row = self.build_average_yields_section(current_row)
        current_row = self.build_machinery_costs_section(current_row)
        current_row = self.build_fuel_section(current_row)
        current_row = self.build_insurance_section(current_row)
        current_row = self.build_tax_rates_section(current_row)
        
        print(f"✓ Settings worksheet created with {len(self.named_ranges)} named ranges")
        
        return self.wb
    
    def save(self, filename="TFOS_Settings_v0.1.xlsx"):
        """Save workbook to file."""
        self.build()
        self.wb.save(filename)
        print(f"✓ Workbook saved: {filename}")
        
        # Print named ranges summary
        print("\nNamed Ranges Created:")
        for item in self.named_ranges:
            print(f"  {item['name']:40} → Settings!${item['cell']:8} ({item['setting']})")
        
        return filename


if __name__ == "__main__":
    builder = SettingsBuilder()
    builder.save("TFOS_Settings_v0.1.xlsx")
    print("\nTFOS Settings worksheet complete.")
